import os
import sys
import zipfile
import datetime
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


def assert_is_valid_xlsx(path: str):
    if not path.lower().endswith(".xlsx"):
        raise ValueError(f"仅支持 .xlsx 文件：{path}")
    if not os.path.exists(path):
        raise FileNotFoundError(f"找不到输入文件：{path}")
    if not zipfile.is_zipfile(path):
        raise ValueError(f"文件不是有效的 xlsx（不是zip结构）：{path}。可能是xls改名或文件损坏。")


def b_sort_key(v):
    """B列排序key：None最后；日期/时间；数值；字符串"""
    if v is None:
        return (1, 0, "")
    if isinstance(v, datetime.datetime):
        return (0, 0, v)
    if isinstance(v, datetime.date):
        return (0, 1, v)
    if isinstance(v, (int, float)):
        return (0, 2, v)
    return (0, 3, str(v))


def to_number(v):
    """把单元格值转成float（失败则0）"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", "").strip()
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def row_is_empty(ws, rr, max_col):
    for c in range(1, max_col + 1):
        v = ws.cell(row=rr, column=c).value
        if v is not None and not (isinstance(v, str) and v.strip() == ""):
            return False
    return True


def snapshot_row(ws, r, max_col):
    """抓取整行：值 + 样式（用于整行移动/复写，保留边框粗细等）"""
    snap = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        snap.append({
            "value": cell.value,
            "style": copy(cell._style),
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
            "comment": cell.comment,
        })
    return snap


def write_snapshot_row(ws, r, snap):
    for c, d in enumerate(snap, start=1):
        cell = ws.cell(row=r, column=c)
        cell.value = d["value"]
        cell._style = copy(d["style"])
        cell.font = copy(d["font"])
        cell.fill = copy(d["fill"])
        cell.border = copy(d["border"])
        cell.alignment = copy(d["alignment"])
        cell.number_format = d["number_format"]
        cell.protection = copy(d["protection"])
        cell.comment = d["comment"]


def copy_dimensions(src_ws, dst_ws):
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden
        dst_ws.column_dimensions[col_letter].outlineLevel = dim.outlineLevel
        dst_ws.column_dimensions[col_letter].collapsed = dim.collapsed


def copy_row_dim(src_ws, dst_ws, src_r, dst_r):
    if src_r in src_ws.row_dimensions:
        sd = src_ws.row_dimensions[src_r]
        dd = dst_ws.row_dimensions[dst_r]
        dd.height = sd.height
        dd.hidden = sd.hidden
        dd.outlineLevel = sd.outlineLevel
        dd.collapsed = sd.collapsed


def fixed_e_value(c_val, d_val, e_val):
    """
    修复E列：如果E为空/0，但C和D有值，则用 E=C*D（保留两位小数）
    """
    c_num = to_number(c_val)
    d_num = to_number(d_val)

    if c_num != 0 and d_num != 0:
        if e_val is None:
            return round(c_num * d_num, 2)
        if isinstance(e_val, (int, float)) and abs(float(e_val)) < 1e-9:
            return round(c_num * d_num, 2)
        if isinstance(e_val, str) and e_val.strip() in {"0", "0.0", "0.00"}:
            return round(c_num * d_num, 2)

    return e_val


def process_excel_xlsx_no_header(input_path: str, output_path: str):
    """
    无表头版：
    - 从第1行开始当数据
    - A列连续相同为一类
    - 每类按B升序排序（整行移动）
    - 先写出排序结果，再计算合计（C/E）
    - 合计行：A/B/D空，C/E合计，B~E黄色
    """
    assert_is_valid_xlsx(input_path)

    wb_formula = load_workbook(input_path, data_only=False)
    wb_values = load_workbook(input_path, data_only=True)
    ws_f = wb_formula.active
    ws_v = wb_values.active

    max_col = ws_f.max_column
    yellow_fill = PatternFill("solid", fgColor="FFFF00")

    start_row = 1  # ✅ 无表头：第一行就是数据

    last = ws_f.max_row
    while last >= start_row and row_is_empty(ws_f, last, max_col):
        last -= 1

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = ws_f.title
    copy_dimensions(ws_f, out_ws)

    out_r = 1
    r = start_row
    
    grand_sum_c = 0.0
    grand_sum_e = 0.0

    while r <= last:
        a_val = ws_f.cell(row=r, column=1).value

        # 空行原样复制
        if a_val is None or (isinstance(a_val, str) and a_val.strip() == ""):
            snap = snapshot_row(ws_f, r, max_col)
            for c in range(1, max_col + 1):
                v_disp = ws_v.cell(row=r, column=c).value
                if v_disp is not None:
                    snap[c - 1]["value"] = v_disp
            write_snapshot_row(out_ws, out_r, snap)
            copy_row_dim(ws_f, out_ws, r, out_r)
            out_r += 1
            r += 1
            continue

        # A连续区块
        start = r
        end = r
        while end + 1 <= last and ws_f.cell(row=end + 1, column=1).value == a_val:
            end += 1

        # 取区块并修复E
        block = []
        for rr in range(start, end + 1):
            snap = snapshot_row(ws_f, rr, max_col)

            # 用显示值覆盖（避免把公式文本写出去）
            for c in range(1, max_col + 1):
                v_disp = ws_v.cell(row=rr, column=c).value
                if v_disp is not None:
                    snap[c - 1]["value"] = v_disp

            if max_col >= 5:
                snap[4]["value"] = fixed_e_value(snap[2]["value"], snap[3]["value"], snap[4]["value"])

            block.append(snap)

        # ✅ 先“出结果”：排序
        block_sorted = sorted(block, key=lambda s: b_sort_key(s[1]["value"] if len(s) >= 2 else None))

        # ✅ 再写结果
        for i, snap in enumerate(block_sorted):
            write_snapshot_row(out_ws, out_r, snap)
            copy_row_dim(ws_f, out_ws, start + i, out_r)
            out_r += 1

        # ✅ 基于最终结果计算合计
        sum_c = sum(to_number(s[2]["value"]) for s in block_sorted) if max_col >= 3 else 0.0
        sum_e = sum(to_number(s[4]["value"]) for s in block_sorted) if max_col >= 5 else 0.0

        grand_sum_c += sum_c
        grand_sum_e += sum_e

        template = snapshot_row(ws_f, end, max_col)
        for c in range(max_col):
            template[c]["value"] = None
        if max_col >= 3:
            template[2]["value"] = sum_c
        if max_col >= 5:
            template[4]["value"] = round(sum_e, 2)

        # B~E 黄色
        for col in range(2, min(5, max_col) + 1):
            template[col - 1]["fill"] = copy(yellow_fill)

        write_snapshot_row(out_ws, out_r, template)
        copy_row_dim(ws_f, out_ws, end, out_r)
        out_r += 1

        r = end + 1

    for merged in ws_f.merged_cells.ranges:
        out_ws.merge_cells(str(merged))
    # ====== 最后追加总合计行（全表，不区分类）======
    template_total = snapshot_row(ws_f, last, max_col)  # 用最后一行当模板保留边框
    for c in range(max_col):
        template_total[c]["value"] = None

    # 建议A列写“总合计”，否则筛选A非空时会看不到
    template_total[0]["value"] = "合计"

    if max_col >= 3:
        template_total[2]["value"] = grand_sum_c
    if max_col >= 5:
        template_total[4]["value"] = round(grand_sum_e, 2)

    # 你想高亮的话：B~E黄色（和小计一致）
    for col in range(2, min(5, max_col) + 1):
        template_total[col - 1]["fill"] = copy(yellow_fill)

    write_snapshot_row(out_ws, out_r, template_total)
    copy_row_dim(ws_f, out_ws, last, out_r)
    out_r += 1

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    out_wb.save(output_path)


def build_default_out_path(in_path: str) -> str:
    base, _ = os.path.splitext(in_path)
    return base + "_处理后.xlsx"


def build_default_out_path(in_path: str) -> str:
    base, _ = os.path.splitext(in_path)
    return base + "_处理后.xlsx"


def gui_pick_file():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()  # 不显示主窗口

    in_path = filedialog.askopenfilename(
        title="选择要处理的 Excel (.xlsx)",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not in_path:
        return None, None, root, messagebox

    out_path = filedialog.asksaveasfilename(
        title="选择输出位置（默认自动命名）",
        defaultextension=".xlsx",
        initialfile=os.path.basename(build_default_out_path(in_path)),
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not out_path:
        # 用户取消保存位置，就用默认输出到同目录
        out_path = build_default_out_path(in_path)

    return in_path, out_path, root, messagebox


def main():
    # 1) 拖拽模式：有参数 -> 直接处理
    if len(sys.argv) >= 2:
        in_path = sys.argv[1].strip('"')
        out_path = sys.argv[2].strip('"') if len(sys.argv) >= 3 else build_default_out_path(in_path)
        try:
            process_excel_xlsx_no_header(in_path, out_path)
            print("完成：", out_path)
        except Exception as e:
            print("处理失败：", str(e))
            raise
        return

    # 2) 双击模式：无参数 -> 弹窗选择文件
    try:
        in_path, out_path, root, messagebox = gui_pick_file()
        if not in_path:
            return

        process_excel_xlsx_no_header(in_path, out_path)
        messagebox.showinfo("完成", f"处理完成！\n输出文件：\n{out_path}")
        root.destroy()
    except Exception as e:
        # 弹窗显示错误
        try:
            import tkinter as tk
            from tkinter import messagebox
            messagebox.showerror("失败", f"处理失败：\n{e}")
        except Exception:
            pass
        raise


if __name__ == "__main__":
    main()
